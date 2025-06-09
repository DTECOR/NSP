"""
Módulo para mostrar la exportación NOC con soporte para múltiples targets.
Permite seleccionar varios equipos y exportar todos sus servicios en un solo archivo.
Optimizado para manejar grandes volúmenes de datos.
"""

import pandas as pd
import streamlit as st
import re
import io
import os
import time
from utils.excel_noc_automatico import cargar_archivos_excel_noc, buscar_servicio_en_ambos_excel, buscar_servicios_en_lote
from utils.formato_excel_profesional import aplicar_formato_profesional_excel
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def mostrar_exportacion_noc(df_servicios, df_resumen):
    """
    Muestra la pestaña de exportación NOC con carga automática o manual de Excel y generación de reportes.
    Soporta selección y exportación de múltiples targets.
    
    Args:
        df_servicios (DataFrame): DataFrame con información de servicios
        df_resumen (DataFrame): DataFrame con el resumen de equipos
    """
    st.header("Exportación NOC")
    
    # Inicializar session_state para almacenar los DataFrames de servicios totales
    if 'df_nsp19' not in st.session_state:
        st.session_state.df_nsp19 = pd.DataFrame()
    
    if 'df_nsp24' not in st.session_state:
        st.session_state.df_nsp24 = pd.DataFrame()
    
    # Inicializar modo de carga en session_state si no existe
    if 'modo_carga_excel' not in st.session_state:
        st.session_state.modo_carga_excel = "automatica"
    
    # Inicializar variables para almacenar datos de exportación
    if 'excel_data_multiple' not in st.session_state:
        st.session_state.excel_data_multiple = None
    
    if 'excel_data_individual' not in st.session_state:
        st.session_state.excel_data_individual = None
    
    if 'estadisticas_multiple' not in st.session_state:
        st.session_state.estadisticas_multiple = None
    
    if 'estadisticas_individual' not in st.session_state:
        st.session_state.estadisticas_individual = None
    
    if 'equipo_individual_seleccionado' not in st.session_state:
        st.session_state.equipo_individual_seleccionado = None
    
    if 'equipos_multiples_seleccionados' not in st.session_state:
        st.session_state.equipos_multiples_seleccionados = []
    
    # Inicializar estado de procesamiento
    if 'procesando_exportacion' not in st.session_state:
        st.session_state.procesando_exportacion = False
    
    if 'exportacion_completada' not in st.session_state:
        st.session_state.exportacion_completada = False
    
    # Selector de método de carga
    st.subheader("Seleccione método de carga:")
    col1, col2 = st.columns(2)
    
    with col1:
        # Radio buttons para seleccionar el método de carga
        modo_carga = st.radio(
            "Método de carga:",
            ["Automática (carpeta InformeNokia)", "Manual (subir archivos)"],
            index=0 if st.session_state.modo_carga_excel == "automatica" else 1,
            key="radio_modo_carga_excel"
        )
        
        # Actualizar el modo de carga en session_state
        st.session_state.modo_carga_excel = "automatica" if modo_carga == "Automática (carpeta InformeNokia)" else "manual"
    
    with col2:
        # Checkbox para cargar archivos automáticamente
        if st.checkbox("Cargar archivos automáticamente", value=True, key="checkbox_cargar_automaticamente"):
            if st.session_state.modo_carga_excel == "automatica":
                # Cargar automáticamente los archivos Excel de servicios
                df_nsp19, df_nsp24 = cargar_archivos_excel_noc()
                st.session_state.df_nsp19 = df_nsp19
                st.session_state.df_nsp24 = df_nsp24
                st.success("Archivos cargados correctamente")
    
    # Mostrar información de los archivos cargados
    col1, col2 = st.columns(2)
    
    with col1:
        if not st.session_state.df_nsp19.empty:
            st.success(f"Archivo NSP19 cargado: {len(st.session_state.df_nsp19)} servicios")
        else:
            st.warning("No se encontró archivo de servicios NSP19")
    
    with col2:
        if not st.session_state.df_nsp24.empty:
            st.success(f"Archivo NSP24 cargado: {len(st.session_state.df_nsp24)} servicios")
        else:
            st.warning("No se encontró archivo de servicios NSP24")
    
    # Si el modo es manual, mostrar los cargadores de archivos directamente
    if st.session_state.modo_carga_excel == "manual":
        st.subheader("Cargar Archivos Manualmente")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.subheader("Archivo de Servicios NSP19")
            uploaded_file_nsp19 = st.file_uploader("Cargar archivo NSP19", type=["xlsx", "csv"], key="exportacion_noc_uploader_nsp19")
            
            if uploaded_file_nsp19 is not None:
                try:
                    # Determinar el tipo de archivo y cargarlo
                    if uploaded_file_nsp19.name.endswith('.csv'):
                        df_nsp19 = pd.read_csv(uploaded_file_nsp19)
                    else:  # Excel
                        df_nsp19 = pd.read_excel(uploaded_file_nsp19)
                    
                    st.session_state.df_nsp19 = df_nsp19
                    st.success(f"Archivo NSP19 cargado correctamente. Se encontraron {len(df_nsp19)} servicios.")
                except Exception as e:
                    st.error(f"Error al cargar el archivo NSP19: {str(e)}")
        
        with col2:
            st.subheader("Archivo de Servicios NSP24")
            uploaded_file_nsp24 = st.file_uploader("Cargar archivo NSP24", type=["xlsx", "csv"], key="exportacion_noc_uploader_nsp24")
            
            if uploaded_file_nsp24 is not None:
                try:
                    # Determinar el tipo de archivo y cargarlo
                    if uploaded_file_nsp24.name.endswith('.csv'):
                        df_nsp24 = pd.read_csv(uploaded_file_nsp24)
                    else:  # Excel
                        df_nsp24 = pd.read_excel(uploaded_file_nsp24)
                    
                    st.session_state.df_nsp24 = df_nsp24
                    st.success(f"Archivo NSP24 cargado correctamente. Se encontraron {len(df_nsp24)} servicios.")
                except Exception as e:
                    st.error(f"Error al cargar el archivo NSP24: {str(e)}")
    
    # Mostrar muestras de los datos
    with st.expander("Ver muestra de datos NSP19"):
        if not st.session_state.df_nsp19.empty:
            st.dataframe(st.session_state.df_nsp19.head(5))
        else:
            st.info("No hay datos disponibles de NSP19")
    
    with st.expander("Ver muestra de datos NSP24"):
        if not st.session_state.df_nsp24.empty:
            st.dataframe(st.session_state.df_nsp24.head(5))
        else:
            st.info("No hay datos disponibles de NSP24")
    
    # Sección para exportar servicios por equipo
    st.subheader("Exportar Servicios para NOC")
    
    # Verificar si hay datos de servicios cargados (al menos uno de los dos)
    if st.session_state.df_nsp19.empty and st.session_state.df_nsp24.empty:
        st.warning("Debe cargar al menos un archivo de servicios (NSP19 o NSP24) para poder exportar.")
    else:
        # Obtener lista de equipos
        equipos = sorted(df_servicios['target'].unique().tolist())
        
        # Crear campo de búsqueda con autocompletado
        equipo_busqueda = st.text_input("Buscar equipo (escriba para filtrar):", "", key="exportacion_noc_busqueda")
        
        # Filtrar equipos según la búsqueda
        equipos_filtrados = [equipo for equipo in equipos if equipo_busqueda.lower() in equipo.lower()]
        
        # Mostrar selector con equipos filtrados
        if equipos_filtrados:
            # Selector para un solo equipo
            st.subheader("Seleccionar un equipo individual:")
            equipo_seleccionado = st.selectbox(
                "Seleccionar Equipo:",
                ['Seleccione un equipo...'] + equipos_filtrados,
                key="equipo_noc_selector_tab"
            )
            
            # Actualizar equipo_noc_seleccionado en session_state
            if equipo_seleccionado != 'Seleccione un equipo...':
                st.session_state.equipo_individual_seleccionado = equipo_seleccionado
                
                # Mostrar servicios del equipo seleccionado
                # Filtrar servicios del equipo seleccionado - CORREGIDO: Filtrado exacto por target
                servicios_equipo = df_servicios[df_servicios['target'].str.strip() == equipo_seleccionado.strip()]
                
                # Mostrar información del equipo
                equipo_info = df_resumen[df_resumen['target'] == equipo_seleccionado]
                if not equipo_info.empty:
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Total Servicios", len(servicios_equipo))
                    with col2:
                        st.metric("Ciudad", equipo_info['ciudad'].values[0])
                    with col3:
                        st.metric("Estado", equipo_info['estado'].values[0])
                
                # Mostrar tabla de servicios
                st.subheader(f"Servicios en {equipo_seleccionado}")
                st.info(f"Se encontraron {len(servicios_equipo)} servicios para este equipo.")
                st.dataframe(servicios_equipo[['service_id', 'type', 'admin_state', 'oper_state', 'customer_id', 'service_name']])
                
                # Botón para exportar a NOC
                if st.button("Exportar a NOC (Equipo Individual)", key="exportar_noc_btn_tab"):
                    try:
                        # Mostrar mensaje de procesamiento
                        with st.spinner("Generando archivo Excel... Por favor espere."):
                            # Generar el archivo Excel para NOC
                            excel_data, estadisticas = generar_excel_noc(
                                servicios_equipo, 
                                st.session_state.df_nsp19, 
                                st.session_state.df_nsp24,
                                equipo_seleccionado
                            )
                            
                            # Guardar en session_state para persistencia
                            st.session_state.excel_data_individual = excel_data
                            st.session_state.estadisticas_individual = estadisticas
                            
                            # Mostrar mensaje de éxito
                            st.success(f"¡Archivo generado correctamente! Haga clic en el botón de descarga para obtener el archivo.")
                    except Exception as e:
                        st.error(f"Error al generar el archivo Excel: {str(e)}")
                        st.info("Detalles técnicos para solución de problemas:")
                        st.code(str(e))
                
                # Mostrar estadísticas y botón de descarga si hay datos generados
                if st.session_state.excel_data_individual is not None and st.session_state.equipo_individual_seleccionado == equipo_seleccionado:
                    # Mostrar estadísticas de origen
                    st.subheader("Origen de los datos")
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("NSP19", st.session_state.estadisticas_individual['NSP19'])
                    with col2:
                        st.metric("NSP24", st.session_state.estadisticas_individual['NSP24'])
                    with col3:
                        st.metric("No encontrados", st.session_state.estadisticas_individual['No encontrado'])
                    
                    # Ofrecer descarga del archivo
                    st.download_button(
                        label="Descargar List of Services",
                        data=st.session_state.excel_data_individual,
                        file_name=f"List of Services - {equipo_seleccionado}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="download_noc_btn_tab"
                    )
            
            # NUEVA SECCIÓN: Selección múltiple de equipos
            st.subheader("Seleccionar múltiples equipos:")
            
            # Opción 1: Multiselect para seleccionar varios equipos
            equipos_seleccionados = st.multiselect(
                "Seleccionar varios equipos:",
                equipos_filtrados,
                key="equipos_noc_multiselect"
            )
            
            # Opción 2: Entrada de texto para lista de equipos
            st.write("O ingrese una lista de equipos (uno por línea):")
            equipos_texto = st.text_area(
                "Lista de equipos (uno por línea):",
                height=100,
                key="equipos_noc_textarea"
            )
            
            # Procesar lista de texto si se proporciona
            todos_equipos_seleccionados = []
            equipos_invalidos = []
            
            if equipos_texto.strip():
                # Dividir por líneas y eliminar espacios en blanco
                equipos_lista = [equipo.strip() for equipo in equipos_texto.split('\n') if equipo.strip()]
                
                # Verificar cuáles existen en el DataFrame
                equipos_validos = [equipo for equipo in equipos_lista if equipo in equipos]
                equipos_invalidos = [equipo for equipo in equipos_lista if equipo not in equipos and equipo.strip()]
                
                # Combinar con los seleccionados en el multiselect
                todos_equipos_seleccionados = list(set(equipos_seleccionados + equipos_validos))
            else:
                todos_equipos_seleccionados = equipos_seleccionados
            
            # Mostrar información sobre equipos inválidos si existen
            if equipos_invalidos:
                st.warning(f"Los siguientes equipos no fueron encontrados: {', '.join(equipos_invalidos)}")
            
            # Mostrar resumen de equipos seleccionados
            if todos_equipos_seleccionados:
                # Guardar en session_state para persistencia
                st.session_state.equipos_multiples_seleccionados = todos_equipos_seleccionados
                
                st.success(f"Total de equipos seleccionados: {len(todos_equipos_seleccionados)}")
                
                # Mostrar lista de equipos seleccionados en un expander
                with st.expander("Ver equipos seleccionados"):
                    st.write(", ".join(todos_equipos_seleccionados))
                
                # Filtrar servicios de todos los equipos seleccionados
                servicios_multiples = df_servicios[df_servicios['target'].isin(todos_equipos_seleccionados)]
                
                # Mostrar información sobre servicios encontrados
                if not servicios_multiples.empty:
                    st.info(f"Se encontraron {len(servicios_multiples)} servicios en total para los equipos seleccionados.")
                    
                    # Botón para exportar servicios de múltiples equipos
                    if st.button("Exportar a NOC (Múltiples Equipos)", key="exportar_noc_multiple_btn"):
                        try:
                            # Activar flag de procesamiento
                            st.session_state.procesando_exportacion = True
                            st.session_state.exportacion_completada = False
                            
                            # Mostrar mensaje de procesamiento
                            with st.spinner(f"Generando archivo Excel para {len(servicios_multiples)} servicios... Este proceso puede tardar varios minutos para grandes volúmenes de datos."):
                                # Generar el archivo Excel para NOC con múltiples equipos
                                excel_data, estadisticas = generar_excel_noc_multiple_optimizado(
                                    servicios_multiples, 
                                    st.session_state.df_nsp19, 
                                    st.session_state.df_nsp24,
                                    todos_equipos_seleccionados
                                )
                                
                                # Guardar en session_state para persistencia
                                st.session_state.excel_data_multiple = excel_data
                                st.session_state.estadisticas_multiple = estadisticas
                                
                                # Actualizar flags
                                st.session_state.procesando_exportacion = False
                                st.session_state.exportacion_completada = True
                                
                                # Mostrar mensaje de éxito
                                st.success(f"¡Archivo generado correctamente! Haga clic en el botón de descarga para obtener el archivo.")
                        except Exception as e:
                            st.session_state.procesando_exportacion = False
                            st.error(f"Error al generar el archivo Excel: {str(e)}")
                            st.info("Detalles técnicos para solución de problemas:")
                            st.code(str(e))
                    
                    # Mostrar estadísticas y botón de descarga si hay datos generados
                    if st.session_state.excel_data_multiple is not None and st.session_state.exportacion_completada:
                        # Verificar si los equipos seleccionados son los mismos que los guardados
                        equipos_actuales = set(todos_equipos_seleccionados)
                        equipos_guardados = set(st.session_state.equipos_multiples_seleccionados)
                        
                        # Mostrar estadísticas de origen
                        st.subheader("Origen de los datos")
                        col1, col2, col3, col4 = st.columns(4)
                        with col1:
                            st.metric("Total Servicios", len(servicios_multiples))
                        with col2:
                            st.metric("NSP19", st.session_state.estadisticas_multiple['NSP19'])
                        with col3:
                            st.metric("NSP24", st.session_state.estadisticas_multiple['NSP24'])
                        with col4:
                            st.metric("No encontrados", st.session_state.estadisticas_multiple['No encontrado'])
                        
                        # Ofrecer descarga del archivo
                        st.download_button(
                            label="Descargar List of Services (Múltiples Equipos)",
                            data=st.session_state.excel_data_multiple,
                            file_name="List of Services - Multiple Targets.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_noc_multiple_btn"
                        )
                else:
                    st.error("No se encontraron servicios para los equipos seleccionados.")
            else:
                st.info("Seleccione al menos un equipo para exportar sus servicios.")
        else:
            st.info("No se encontraron equipos que coincidan con la búsqueda.")

def generar_excel_noc(servicios_equipo, df_nsp19, df_nsp24, nombre_equipo):
    """
    Genera un archivo Excel con el formato requerido para el NOC.
    Busca en ambos archivos NSP19 y NSP24, e informa el origen de los datos.
    
    Args:
        servicios_equipo (DataFrame): DataFrame con los servicios del equipo seleccionado
        df_nsp19 (DataFrame): DataFrame con servicios de NSP19
        df_nsp24 (DataFrame): DataFrame con servicios de NSP24
        nombre_equipo (str): Nombre del equipo seleccionado
    
    Returns:
        tuple: (bytes, dict) Datos del archivo Excel generado y estadísticas de origen
    """
    # Crear DataFrame para el reporte NOC
    df_noc = pd.DataFrame(columns=['Service ID', 'Customer/Company', 'Name', 'Service Impact', 'Origen'])
    
    # Estadísticas de origen
    estadisticas = {
        'NSP19': 0,
        'NSP24': 0,
        'No encontrado': 0
    }
    
    # Procesar cada servicio
    for _, servicio in servicios_equipo.iterrows():
        # Obtener el ServiceId exacto
        service_id = str(servicio['service_id']).strip()
        
        # Obtener el nombre del servicio del archivo original (puede estar truncado)
        service_name_original = servicio['service_name'] if pd.notna(servicio['service_name']) else ""
        
        # Buscar la descripción completa en ambos Excel/CSV usando el ServiceId exacto
        descripcion_completa, origen = buscar_servicio_en_ambos_excel(service_id, df_nsp19, df_nsp24)
        
        # Actualizar estadísticas
        estadisticas[origen] += 1
        
        # Si encontramos la descripción completa en algún Excel/CSV, usarla; si no, usar la del archivo original
        nombre_completo = descripcion_completa if descripcion_completa is not None else service_name_original
        
        # Extraer el código CI o CO de la descripción completa usando la nueva lógica mejorada
        id_tipo = extraer_codigo_ci_co_mejorado(service_id, nombre_completo)
        
        # Añadir fila al DataFrame NOC
        df_noc = pd.concat([df_noc, pd.DataFrame({
            'Service ID': [id_tipo],  # El código CI o CO extraído con la nueva lógica
            'Customer/Company': ['LibertyNet'],  # Siempre es LibertyNet
            'Name': [nombre_completo],  # La descripción completa del servicio
            'Service Impact': ['Loss of Service'],  # Siempre es Loss of Service
            'Origen': [origen]  # NSP19, NSP24 o No encontrado
        })], ignore_index=True)
    
    # Crear una copia del DataFrame sin la columna de origen para exportar
    df_noc_export = df_noc.drop(columns=['Origen'])
    
    # Aplicar formato profesional al Excel
    output = aplicar_formato_profesional_excel(df_noc_export, nombre_equipo)
    
    # Devolver los datos del archivo y las estadísticas
    return output.getvalue(), estadisticas

def generar_excel_noc_multiple_optimizado(servicios_multiples, df_nsp19, df_nsp24, equipos_seleccionados):
    """
    Genera un archivo Excel con el formato requerido para el NOC para múltiples equipos.
    Versión optimizada para manejar grandes volúmenes de datos.
    
    Args:
        servicios_multiples (DataFrame): DataFrame con los servicios de todos los equipos seleccionados
        df_nsp19 (DataFrame): DataFrame con servicios de NSP19
        df_nsp24 (DataFrame): DataFrame con servicios de NSP24
        equipos_seleccionados (list): Lista de nombres de equipos seleccionados
    
    Returns:
        tuple: (bytes, dict) Datos del archivo Excel generado y estadísticas de origen
    """
    # Usar la función optimizada para procesar servicios en lotes
    from utils.excel_noc_automatico import buscar_servicios_en_lote
    
    # Procesar servicios en lotes
    resultados, estadisticas = buscar_servicios_en_lote(servicios_multiples, df_nsp19, df_nsp24)
    
    # Crear DataFrame a partir de los resultados
    df_noc = pd.DataFrame(resultados)
    
    # Aplicar la nueva lógica mejorada de extracción de Service ID a todos los registros
    df_noc['Service ID'] = df_noc.apply(
        lambda row: extraer_codigo_ci_co_mejorado(row.get('Service ID', ''), row.get('Name', '')), 
        axis=1
    )
    
    # Crear una copia del DataFrame sin la columna de origen para exportar
    df_noc_export = df_noc.drop(columns=['Origen'])
    
    # Aplicar formato profesional al Excel para múltiples equipos
    output = aplicar_formato_profesional_excel_multiple(df_noc_export, equipos_seleccionados)
    
    # Devolver los datos del archivo y las estadísticas
    return output.getvalue(), estadisticas

def extraer_codigo_ci_co_mejorado(service_id, descripcion):
    """
    Extrae el código CI o CO de la descripción del servicio con reglas mejoradas.
    
    Reglas:
    1. Si ya tiene formato CI/CO explícito, lo mantiene
    2. Si contiene un número de 6+ dígitos, añade prefijo CI
    3. Si no cumple ninguna condición, devuelve "Sin ID"
    
    Args:
        service_id (str): ID del servicio original
        descripcion (str): Descripción completa del servicio
        
    Returns:
        str: Código CI o CO extraído según las reglas, o "Sin ID" si no se encuentra
    """
    # Verificar entradas válidas
    if not service_id or not isinstance(service_id, str):
        service_id = ""
    
    if not descripcion or not isinstance(descripcion, str):
        descripcion = ""
    
    # Convertir a string y eliminar espacios
    service_id = str(service_id).strip()
    descripcion = str(descripcion).strip()
    
    # REGLA 1: Verificar si ya tiene formato CI/CO explícito
    # Patrones para buscar códigos CI o CO explícitos
    patrones_explicitos = [
        r'CI\d{6,}',          # CI seguido de 6+ dígitos
        r'CO\d{6,}',          # CO seguido de 6+ dígitos
        r'CI\d{6,}[_-]',      # CI seguido de 6+ dígitos y guion/underscore
        r'CO\d{6,}[_-]',      # CO seguido de 6+ dígitos y guion/underscore
    ]
    
    # Buscar en service_id primero
    for patron in patrones_explicitos:
        match = re.search(patron, service_id, re.IGNORECASE)
        if match:
            # Extraer solo la parte CI/CO + números
            return re.search(r'(CI|CO)\d{6,}', match.group(0), re.IGNORECASE).group(0).upper()
    
    # Luego buscar en descripción
    for patron in patrones_explicitos:
        match = re.search(patron, descripcion, re.IGNORECASE)
        if match:
            # Extraer solo la parte CI/CO + números
            return re.search(r'(CI|CO)\d{6,}', match.group(0), re.IGNORECASE).group(0).upper()
    
    # REGLA 2: Buscar números de 6+ dígitos para añadir prefijo CI
    # Primero en service_id
    numero_match = re.search(r'\d{6,}', service_id)
    if numero_match:
        return f"CI{numero_match.group(0)}"
    
    # Luego en descripción
    numero_match = re.search(r'\d{6,}', descripcion)
    if numero_match:
        return f"CI{numero_match.group(0)}"
    
    # REGLA 3: Casos especiales que no deben considerarse como IDs válidos
    palabras_clave_invalidas = ['COM', 'VPLS', 'EPIPE', '_tmnx_', 'GESTION']
    
    for palabra in palabras_clave_invalidas:
        if palabra.lower() in service_id.lower() or palabra.lower() in descripcion.lower():
            return "Sin ID"
    
    # Si service_id es muy corto o no tiene formato numérico adecuado
    if len(service_id) < 6 or not re.search(r'\d{3,}', service_id):
        return "Sin ID"
    
    # Si llegamos aquí, no se encontró un ID válido
    return "Sin ID"

def aplicar_formato_profesional_excel_multiple(df, equipos_seleccionados):
    """
    Aplica un formato profesional al DataFrame para exportación a Excel con múltiples equipos.
    
    Args:
        df (DataFrame): DataFrame con los datos a exportar
        equipos_seleccionados (list): Lista de nombres de equipos seleccionados
        
    Returns:
        BytesIO: Objeto de memoria con el Excel formateado
    """
    # Crear un objeto BytesIO para guardar el Excel en memoria
    output = io.BytesIO()
    
    # Crear un ExcelWriter con openpyxl como motor
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Guardar el DataFrame sin índice
        df.to_excel(writer, sheet_name='List of Services', index=False, startrow=2)  # Empezar en fila 3 para dejar espacio para el título
        
        # Obtener el libro y la hoja
        workbook = writer.book
        worksheet = writer.sheets['List of Services']
        
        # Definir estilos
        naranja_header = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
        borde_fino = Side(style='thin', color='000000')
        borde_completo = Border(left=borde_fino, right=borde_fino, top=borde_fino, bottom=borde_fino)
        
        # Añadir título con el número de equipos
        worksheet.merge_cells('A1:E1')
        titulo_celda = worksheet['A1']
        titulo_celda.value = f"Servicios de {len(equipos_seleccionados)} equipos seleccionados"
        titulo_celda.font = Font(bold=True, size=14)
        titulo_celda.alignment = Alignment(horizontal='center', vertical='center')
        
        # Formatear encabezados (fila 3, que es la 4 en Excel)
        for col_num, column_title in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_num)
            celda = worksheet[f'{col_letter}3']
            celda.fill = naranja_header
            celda.font = Font(bold=True, color='FFFFFF')
            celda.alignment = Alignment(horizontal='center', vertical='center')
            celda.border = borde_completo
            
            # Ajustar ancho de columna
            worksheet.column_dimensions[col_letter].width = max(15, len(str(column_title)) + 5)
        
        # Aplicar bordes y alineación a todas las celdas con datos
        for row in range(4, len(df) + 4):  # +4 porque empezamos en la fila 3 y hay que contar desde 1
            for col_num in range(1, len(df.columns) + 1):
                col_letter = get_column_letter(col_num)
                celda = worksheet[f'{col_letter}{row}']
                celda.border = borde_completo
                
                # Centrar columnas específicas
                if col_num in [1, 3, 5]:  # Target, Customer/Company y Service Impact
                    celda.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ajustar altura de filas
        worksheet.row_dimensions[1].height = 30  # Título
        worksheet.row_dimensions[3].height = 20  # Encabezados
        
        # Crear una hoja adicional con la lista de equipos
        equipos_df = pd.DataFrame({'Equipo': equipos_seleccionados})
        equipos_df.to_excel(writer, sheet_name='Equipos', index=False, startrow=1)
        
        # Formatear hoja de equipos
        equipos_sheet = writer.sheets['Equipos']
        
        # Título para la hoja de equipos
        equipos_sheet.merge_cells('A1:A1')
        equipos_titulo = equipos_sheet['A1']
        equipos_titulo.value = "Lista de Equipos Seleccionados"
        equipos_titulo.font = Font(bold=True, size=12)
        equipos_titulo.alignment = Alignment(horizontal='center')
        
        # Formatear encabezado
        celda = equipos_sheet['A2']
        celda.fill = naranja_header
        celda.font = Font(bold=True, color='FFFFFF')
        celda.alignment = Alignment(horizontal='center')
        celda.border = borde_completo
        
        # Ajustar ancho de columna
        equipos_sheet.column_dimensions['A'].width = 30
        
        # Aplicar bordes a todas las celdas con datos
        for row in range(3, len(equipos_seleccionados) + 3):
            celda = equipos_sheet[f'A{row}']
            celda.border = borde_completo
    
    # Regresar al inicio del stream
    output.seek(0)
    return output
