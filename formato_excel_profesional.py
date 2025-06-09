"""
Módulo para mejorar el formato del Excel exportado para NOC.
Implementa un estilo profesional con encabezados naranjas y línea de título con el equipo.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

def aplicar_formato_profesional_excel(df, nombre_equipo):
    """
    Aplica un formato profesional al DataFrame para exportación a Excel.
    
    Args:
        df (DataFrame): DataFrame con los datos a exportar
        nombre_equipo (str): Nombre del equipo para el título
        
    Returns:
        BytesIO: Objeto de memoria con el Excel formateado
    """
    # Crear un objeto BytesIO para guardar el Excel en memoria
    output = BytesIO()
    
    # Crear un ExcelWriter con openpyxl como motor
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Guardar el DataFrame sin índice
        df.to_excel(writer, sheet_name='List of Services', index=False, startrow=2)  # Empezar en fila 3 para dejar espacio para el título
        
        # Obtener el libro y la hoja
        workbook = writer.book
        worksheet = writer.sheets['List of Services']
        
        # Definir estilos
        naranja_header = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
        borde_fino = Side(style='thin', color='000000')
        borde_completo = Border(left=borde_fino, right=borde_fino, top=borde_fino, bottom=borde_fino)
        
        # Añadir título con el nombre del equipo
        worksheet.merge_cells('A1:D1')
        titulo_celda = worksheet['A1']
        titulo_celda.value = f"Servicios del equipo: {nombre_equipo}"
        titulo_celda.font = Font(bold=True, size=14)
        titulo_celda.alignment = Alignment(horizontal='center', vertical='center')
        
        # Formatear encabezados (fila 3, que es la 4 en Excel)
        for col_num, column_title in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_num)
            celda = worksheet[f'{col_letter}3']
            celda.fill = naranja_header
            celda.font = Font(bold=True, color='FFFFFF')
            celda.alignment = Alignment(horizontal='center', vertical='center')
            celda.border = borde_completo
            
            # Ajustar ancho de columna
            worksheet.column_dimensions[col_letter].width = max(15, len(str(column_title)) + 5)
        
        # Aplicar bordes y alineación a todas las celdas con datos
        for row in range(4, len(df) + 4):  # +4 porque empezamos en la fila 3 y hay que contar desde 1
            for col_num in range(1, len(df.columns) + 1):
                col_letter = get_column_letter(col_num)
                celda = worksheet[f'{col_letter}{row}']
                celda.border = borde_completo
                
                # Centrar columnas específicas
                if col_num in [2, 4]:  # Customer/Company y Service Impact
                    celda.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ajustar altura de filas
        worksheet.row_dimensions[1].height = 30  # Título
        worksheet.row_dimensions[3].height = 20  # Encabezados
        
        # Crear una hoja adicional con estadísticas de origen si existe la columna 'Origen'
        if 'Origen' in df.columns:
            # Crear DataFrame de estadísticas
            estadisticas = df['Origen'].value_counts().reset_index()
            estadisticas.columns = ['Origen', 'Cantidad']
            
            # Guardar estadísticas en una hoja separada
            estadisticas.to_excel(writer, sheet_name='Estadísticas', index=False, startrow=1)
            
            # Formatear hoja de estadísticas
            stats_sheet = writer.sheets['Estadísticas']
            
            # Título para la hoja de estadísticas
            stats_sheet.merge_cells('A1:B1')
            stats_titulo = stats_sheet['A1']
            stats_titulo.value = "Estadísticas de origen de datos"
            stats_titulo.font = Font(bold=True, size=12)
            stats_titulo.alignment = Alignment(horizontal='center')
            
            # Formatear encabezados
            for col_num, column_title in enumerate(['Origen', 'Cantidad'], 1):
                col_letter = get_column_letter(col_num)
                celda = stats_sheet[f'{col_letter}2']
                celda.fill = naranja_header
                celda.font = Font(bold=True, color='FFFFFF')
                celda.alignment = Alignment(horizontal='center')
                celda.border = borde_completo
                
                # Ajustar ancho de columna
                stats_sheet.column_dimensions[col_letter].width = max(15, len(column_title) + 5)
            
            # Aplicar bordes y alineación a todas las celdas con datos
            for row in range(3, len(estadisticas) + 3):
                for col_num in range(1, 3):
                    col_letter = get_column_letter(col_num)
                    celda = stats_sheet[f'{col_letter}{row}']
                    celda.border = borde_completo
                    if col_num == 2:  # Columna de cantidad
                        celda.alignment = Alignment(horizontal='center')
    
    # Regresar al inicio del stream
    output.seek(0)
    return output
